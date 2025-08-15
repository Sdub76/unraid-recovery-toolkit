# Reminder: before running, activate your venv (e.g. `source $HOME/recovery-venv/bin/activate`)


#!/usr/bin/env python3
# path_levels.py — Summarize file counts by directory depth (N+ roll-up) into an Excel workbook.
#
# Features
# - N+ counting: every file contributes once to each requested level.
# - Component-boundary filter: --folder matches exactly that folder and below (case-sensitive).
# - Progress bar with tqdm (updates at least every 5 seconds).
# - Output: path_levels_YYYYMMDD_HHMMSS.xlsx in the current working directory.
# - Sheets: level1..levelN with columns [Count, Path], alphabetically sorted by Path, totals row at bottom.
# - Root files (no directory components) are grouped under the bucket "(root)".
#
# Usage
#     python path_levels.py [--levels N] [--folder FOLDER] <input_file>
#
# Examples
#     python path_levels.py filelist.disk8.txt
#     python path_levels.py --levels 4 --folder tv/Library/Ken filelist.disk8.txt

import argparse
import os
import sys

def load_backup_folders(file_path="backup_folders.txt"):
    if not os.path.isfile(file_path):
        print(f"Backup folder list not found: {file_path}", file=sys.stderr)
        sys.exit(2)
    folders = set()
    with open(file_path, "r", encoding="utf-8") as f:
        for line in f:
            s = line.strip()
            if not s or s.startswith("#"):
                continue
            folders.add(s)
    return folders


from collections import defaultdict
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Font  # pip install openpyxl
from tqdm import tqdm  # pip install tqdm


ROOT_BUCKET = "(root)"


def count_lines_binary(file_path: str, chunk_size: int = 8 * 1024 * 1024) -> int:
    """
    Fast line count in binary mode to avoid decode errors in the counting pass.
    Returns the number of newline characters. Handles files without a trailing newline.
    """
    total = 0
    with open(file_path, "rb") as f:
        while True:
            chunk = f.read(chunk_size)
            if not chunk:
                break
            total += chunk.count(b"\n")
    # If the file does not end with newline but is non-empty, count the last line
    try:
        size = os.path.getsize(file_path)
        if size > 0:
            with open(file_path, "rb") as f:
                f.seek(-1, os.SEEK_END)
                last_char = f.read(1)
                if last_char != b"\n":
                    total += 1
    except Exception:
        # If we can't probe the last byte safely, ignore—best-effort count.
        pass
    return total


def parse_args() -> argparse.Namespace:
    p = argparse.ArgumentParser(
        description="Summarize file counts by directory depth (N+ roll-up) into an Excel workbook."
    )
    p.add_argument("--levels", type=int, default=1, help="Number of depth levels to summarize (default: 1).")
    p.add_argument(
        "--folder",
        type=str,
        default=None,
        help=(
            "Case-sensitive component-boundary prefix to filter paths. "
            "Matches the folder itself or any descendant (e.g., 'tv/Library/Ken' matches 'tv/Library/Ken/...')."
        ),
    )
    p.add_argument("--backup-file", type=str, default="backup_folders.txt", help="Path to backup_folders.txt (default: backup_folders.txt).")
    p.add_argument("input_file", type=str, help="Path to the input file list (UTF-8, one path per line).")
    return p.parse_args()


def validate_args(args: argparse.Namespace) -> None:
    if args.levels < 1:
        print("--levels must be >= 1", file=sys.stderr)
        sys.exit(2)
    if not os.path.isfile(args.input_file):
        print(f"Input file not found: {args.input_file}", file=sys.stderr)
        sys.exit(2)


def folder_matcher(folder: str):
    """
    Return a predicate that checks path-component boundary match:
    - path == folder (rare for files-only lists)
    - path starts with folder + '/'
    Case-sensitive, Unix-style.
    """
    if folder is None:
        return lambda p: True
    # Normalize: strip trailing slashes on folder (keep case)
    while folder.endswith("/"):
        folder = folder[:-1]
    prefix = folder + "/"
    def _match(path: str) -> bool:
        return path == folder or path.startswith(prefix)
    return _match


def bucket_keys_for_path(path: str, levels: int):
    """
    Given a file path like 'a/b/c/file.ext', return the N+ bucket keys for N in [1..levels].
    - Use only directory components (exclude the filename).
    - If the path has fewer than N directory components, use the deepest directory (or ROOT_BUCKET if none).
    Examples:
      path='a/b/c/file.ext', levels=4
        dir_parts = ['a','b','c']
        level1 -> 'a'
        level2 -> 'a/b'
        level3 -> 'a/b/c'
        level4 -> 'a/b/c' (stays at deepest dir, since there is no 4th dir)
      path='file.ext', levels=3
        dir_parts = []
        level1 -> '(root)'
        level2 -> '(root)'
        level3 -> '(root)'
    """
    parts = path.split("/")
    if not parts:
        return [ROOT_BUCKET] * levels
    dir_parts = parts[:-1]  # drop the filename
    keys = []
    if len(dir_parts) == 0:
        keys = [ROOT_BUCKET] * levels
    else:
        for n in range(1, levels + 1):
            if len(dir_parts) >= n:
                keys.append("/".join(dir_parts[:n]))
            else:
                keys.append("/".join(dir_parts))
    return keys


def main() -> None:
    args = parse_args()
    validate_args(args)

    # Pre-count lines for a nice progress bar
    total_lines = count_lines_binary(args.input_file)
    if total_lines == 0:
        print("Input file is empty.", file=sys.stderr)
        sys.exit(2)

    match = folder_matcher(args.folder)

    # One counter per level
    level_counters = [defaultdict(int) for _ in range(args.levels)]

    processed = 0
    filtered_out = 0
    blank_lines = 0

    # Fail hard on decode errors, trim whitespace (but don't alter internal spaces)
    with open(args.input_file, "r", encoding="utf-8", errors="strict") as f, tqdm(
        total=total_lines, unit=" lines", mininterval=5.0, desc="Processing"
    ) as bar:
        for line in f:
            bar.update(1)
            s = line.strip()
            if not s:
                blank_lines += 1
                continue

            # Folder filter (component-boundary)
            if not match(s):
                filtered_out += 1
                continue

            processed += 1
            keys = bucket_keys_for_path(s, args.levels)
            for idx, key in enumerate(keys):
                level_counters[idx][key] += 1

    if processed == 0:
        msg = "No matching records to summarize"
        if args.folder:
            msg += f" (check --folder='{args.folder}' ?)"
        print(msg, file=sys.stderr)
        sys.exit(2)

    # Prepare Excel workbook
    wb = Workbook()
    # By default, openpyxl creates one sheet named 'Sheet'; we'll replace it
    default_ws = wb.active
    wb.remove(default_ws)

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    in_stem = os.path.splitext(os.path.basename(args.input_file))[0]
    
    # Ensure output directory exists
    os.makedirs("out", exist_ok=True)
    out_name = f"out/{in_stem}_{timestamp}.xlsx"

    
    
    # Write sheets Level 1..Level N
    backup_set = load_backup_folders(args.backup_file)
    for i, counter in enumerate(level_counters, start=1):
        ws = wb.create_sheet(title=f"Level {i}")
        # Header: Backup + Count + Path 1..Path i
        header = ["Backup", "Count"] + [f"Path {k}" for k in range(1, i + 1)]
        ws.append(header)

        # Sort alphabetically by full key/path (string)
        items = sorted(counter.items(), key=lambda kv: kv[0])

        total = 0
        for key, count in items:
            if key == ROOT_BUCKET:
                parts = []
            else:
                parts = key.split("/")

            # Determine if top-level folder is in backup set
            top_level = parts[0] if parts else ROOT_BUCKET
            backup_flag = "true" if top_level in backup_set else ""

            row = [backup_flag, count]
            for idx in range(i):
                if idx < len(parts):
                    row.append(parts[idx])
                else:
                    if key == ROOT_BUCKET and idx == 0:
                        row.append(ROOT_BUCKET)
                    else:
                        row.append("")
            ws.append(row)
            total += count

        # Totals row
        total_row = ["", total] + ["TOTAL"] + [""] * (i - 1)
        ws.append(total_row)
        for cell in ws[ws.max_row]:
            cell.font = Font(bold=True)

    # Save workbook
    wb.save(out_name)



    # Console summary
    print("\n=== Summary ===")
    print(f"Input file: {args.input_file}")
    if args.folder:
        print(f"Filter:     {args.folder} (component-boundary, case-sensitive)")
    print(f"Levels:     {args.levels}")
    print(f"Processed:  {processed:,} files")
    if filtered_out:
        print(f"Filtered:   {filtered_out:,} lines (did not match --folder)")
    if blank_lines:
        print(f"Blank:      {blank_lines:,} lines ignored")

    for i, counter in enumerate(level_counters, start=1):
        distinct = len(counter)
        print(f"\nLevel {i}: {distinct:,} buckets")
        # Top 10 by count desc, then path asc
        top10 = sorted(counter.items(), key=lambda kv: (-kv[1], kv[0]))[:10]
        for path_key, count in top10:
            print(f"  {count:>10}  {path_key}")

    print(f"\nWrote: {out_name}")


if __name__ == "__main__":
    main()