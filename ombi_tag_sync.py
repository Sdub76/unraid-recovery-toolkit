#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
Ombi → Radarr/Sonarr tag backfill to multi-sheet XLSX (no status filtering).
- TV requester extracted from childRequests
- Root folder lookup (Radarr/Sonarr) + path prefix inference
- Movie year fuzz (±N, default 1) when matching (title,year)
- Dedup: keep earliest request per (requester, ID/title) + write duplicates to extra tabs
- Per-user request summary printed to console
- OUTPUT: one XLSX with 6 tabs:
    - radarr_preview      requestDate,requester,tmdbId,radarrId,title,rootFolder,note
    - sonarr_preview      requestDate,requester,tvdbId,sonarrId,title,rootFolder,note
    - radarr_unmatched    requestDate,requester,tmdbId,radarrId,title,rootFolder,note
    - sonarr_unmatched    requestDate,requester,tvdbId,sonarrId,title,rootFolder,note
    - radarr_duplicates_removed   (same columns as preview)
    - sonarr_duplicates_removed   (same columns as preview)

ENV VARS (required):
  OMBI_URL, OMBI_API_KEY
  RADARR_URL, RADARR_API_KEY
  SONARR_URL, SONARR_API_KEY

Default: preview only (writes XLSX).
--write radarr      apply tags to Radarr movies
--write sonarr      apply tags to Sonarr series
--apply-tags MODE   add|replace|remove (default: add)
--out               output .xlsx file path (default ./ombi_tag_preview.xlsx)
--year-fuzz N       year ±N for title-year matching (default 1)
"""

import argparse, os, sys, time, requests
from collections import Counter
from typing import Dict, List, Optional, Tuple
from datetime import datetime

try:
    from openpyxl import Workbook
    from openpyxl.utils import get_column_letter
except Exception as e:
    print("This script requires 'openpyxl'. Install it with:\n  pip install openpyxl", flush=True)
    raise

# ---------- tiny progress/err helpers ----------
def progress(msg: str) -> None:
    print(msg, flush=True)

def backoff_sleep(attempt: int, base: float) -> None:
    time.sleep(base * (attempt + 1))

# ---------- HTTP with retries ----------
class Http:
    def __init__(self, timeout: int = 15, max_retries: int = 3, backoff: float = 1.5):
        self.s = requests.Session(); self.timeout=timeout; self.max_retries=max_retries; self.backoff=backoff
    def _req(self, method, url, **kw):
        last = None
        for i in range(self.max_retries):
            try:
                r = self.s.request(method, url, timeout=self.timeout, **kw)
                if r.status_code >= 500:
                    raise requests.HTTPError(f"{r.status_code} server error: {r.text[:200]}")
                return r
            except Exception as e:
                last = e
                progress(f"[HTTP] {method} {url} failed attempt {i+1}: {e}")
                backoff_sleep(i, self.backoff)
        raise last
    def get(self, url, **kw):  return self._req("GET", url, **kw)
    def post(self, url, **kw): return self._req("POST", url, **kw)
    def put(self, url, **kw):  return self._req("PUT", url, **kw)

# ---------- Ombi ----------
class Ombi:
    def __init__(self, base: str, key: str, http: Http):
        self.base = base.rstrip("/"); self.h = {"ApiKey": key, "Content-Type":"application/json"}; self.http=http

    @staticmethod
    def _pick_request_date(obj: dict) -> str:
        for k in ("requestedDate", "requestedDateUtc", "requestedAt", "createdAt", "requestDate", "dateRequested"):
            v = obj.get(k)
            if v:
                return str(v)
        return ""

    @staticmethod
    def _pick_requester_flat_or_nested(obj: dict) -> str:
        ru = obj.get("requestedUser") or {}
        for k in ("userName", "alias", "username", "name", "displayName", "userAlias"):
            v = ru.get(k)
            if v:
                return str(v)
        for k in (
            "requestedUserName", "requestedByName", "requestedBy",
            "userAlias", "userName", "username", "alias", "user"
        ):
            v = obj.get(k)
            if v:
                return str(v)
        rq = obj.get("request") or {}
        for k in ("requestedUserName", "requestedByName", "requestedBy", "userName", "username", "alias"):
            v = rq.get(k)
            if v:
                return str(v)
        return ""

    def _rest_movies(self) -> Optional[List[dict]]:
        url = f"{self.base}/api/v1/Request/movie"
        try:
            r = self.http.get(url, headers=self.h)
            if r.status_code == 404:
                return None
            r.raise_for_status()
            data = r.json()
            out = []
            for it in data:
                requester = self._pick_requester_flat_or_nested(it)
                title = it.get("title") or (it.get("theMovieDb") or {}).get("title") or ""
                year  = it.get("year") or it.get("releaseYear") or None
                tmdb  = it.get("theMovieDbId") or it.get("tmdbId")
                if isinstance(tmdb, dict): tmdb = tmdb.get("id") or tmdb.get("theMovieDbId")
                request_date = self._pick_request_date(it) or self._pick_request_date(it.get("request") or {})
                out.append({
                    "requester": requester,
                    "title": title,
                    "year": year,
                    "tmdbId": tmdb if isinstance(tmdb,int) else None,
                    "requestDate": request_date,
                })
            return out
        except Exception as e:
            progress(f"[Ombi] REST fetch failed /Request/movie: {e}")
            return None

    def _rest_tv(self) -> Optional[List[dict]]:
        """Flatten TV into one row per (series, childRequest/requester)."""
        url = f"{self.base}/api/v1/Request/tv"
        try:
            r = self.http.get(url, headers=self.h)
            if r.status_code == 404:
                return None
            r.raise_for_status()
            shows = r.json()
            out: List[dict] = []
            for show in shows:
                title = show.get("title") or show.get("seriesName") or ""
                tvdb  = show.get("tvDbId") or show.get("tvdbId")
                child_list = show.get("childRequests") or []
                if not child_list:
                    out.append({
                        "title": title,
                        "tvdbId": tvdb if isinstance(tvdb,int) else None,
                        "requester": self._pick_requester_flat_or_nested(show),
                        "requestDate": self._pick_request_date(show),
                    })
                    continue
                for cr in child_list:
                    requester = self._pick_requester_flat_or_nested(cr) or self._pick_requester_flat_or_nested(show)
                    request_date = self._pick_request_date(cr) or self._pick_request_date(show)
                    out.append({
                        "title": title,
                        "tvdbId": tvdb if isinstance(tvdb,int) else None,
                        "requester": requester or "",
                        "requestDate": request_date or "",
                    })
            return out
        except Exception as e:
            progress(f"[Ombi] REST fetch failed /Request/tv: {e}")
            return None

    def fetch_all(self) -> Tuple[List[dict], List[dict]]:
        progress("Step 1/9: Fetching Ombi requests (movies & tv)…")
        m = self._rest_movies(); t = self._rest_tv()
        if m is None: m = []
        if t is None: t = []
        progress(f"Ombi items: movies={len(m)} tv-requests={len(t)}")
        return m, t

# ---------- Radarr / Sonarr ----------
class Radarr:
    def __init__(self, base: str, key: str, http: Http):
        self.base = base.rstrip("/"); self.h={"X-Api-Key": key}; self.http=http
    def all_movies(self) -> List[dict]:
        progress("Step 2/9: Loading Radarr library…")
        r = self.http.get(f"{self.base}/api/v3/movie", headers=self.h); r.raise_for_status()
        data = r.json(); progress(f"Radarr movies: {len(data)}"); return data
    def root_folders(self) -> List[dict]:
        r = self.http.get(f"{self.base}/api/v3/rootfolder", headers=self.h); r.raise_for_status()
        return r.json()
    def get_or_create_tag(self, label: str) -> int:
        r = self.http.get(f"{self.base}/api/v3/tag", headers=self.h); r.raise_for_status()
        for t in r.json():
            if t.get("label","") == label: return t["id"]
        r = self.http.post(f"{self.base}/api/v3/tag", headers=self.h, json={"label": label}); r.raise_for_status()
        return r.json()["id"]
    def apply_tag(self, ids: List[int], tag_id: int, mode: str):
        if not ids: return
        payload = {"movieIds": sorted(set(ids)), "tags":[tag_id], "applyTags": mode}
        r = self.http.put(f"{self.base}/api/v3/movie/editor", headers=self.h, json=payload); r.raise_for_status()

class Sonarr:
    def __init__(self, base: str, key: str, http: Http):
        self.base = base.rstrip("/"); self.h={"X-Api-Key": key}; self.http=http
    def all_series(self) -> List[dict]:
        progress("Step 3/9: Loading Sonarr library…")
        r = self.http.get(f"{self.base}/api/v3/series", headers=self.h); r.raise_for_status()
        data = r.json(); progress(f"Sonarr series: {len(data)}"); return data
    def root_folders(self) -> List[dict]:
        r = self.http.get(f"{self.base}/api/v3/rootfolder", headers=self.h); r.raise_for_status()
        return r.json()
    def get_or_create_tag(self, label: str) -> int:
        r = self.http.get(f"{self.base}/api/v3/tag", headers=self.h); r.raise_for_status()
        for t in r.json():
            if t.get("label","") == label: return t["id"]
        r = self.http.post(f"{self.base}/api/v3/tag", headers=self.h, json={"label": label}); r.raise_for_status()
        return r.json()["id"]
    def apply_tag(self, ids: List[int], tag_id: int, mode: str):
        if not ids: return
        payload = {"seriesIds": sorted(set(ids)), "tags":[tag_id], "applyTags": mode}
        r = self.http.put(f"{self.base}/api/v3/series/editor", headers=self.h, json=payload); r.raise_for_status()

# ---------- mapping helpers ----------
def build_radarr_maps(movies: List[dict], rad_rootfolders: List[dict]):
    progress("Step 4/9: Building Radarr ID maps…")
    tmdb_to_id, ty_to_id = {}, {}
    id_to_root = {}
    roots = [rf.get("path") or rf.get("rootFolderPath") or "" for rf in (rad_rootfolders or [])]
    roots = sorted({p.rstrip("/\\") for p in roots if p}, key=len, reverse=True)
    for m in movies:
        mid = m.get("id")
        tmdb = m.get("tmdbId")
        title=(m.get("title") or "").strip().lower()
        year=m.get("year")
        if isinstance(tmdb,int): tmdb_to_id[tmdb]=mid
        if title and isinstance(year,int): ty_to_id[(title,year)]=mid
        rf = m.get("rootFolderPath") or ""
        if not rf:
            path = (m.get("path") or "").rstrip("/\\")
            rf = ""
            for r in roots:
                if path.startswith(r):
                    rf = r
                    break
        id_to_root[mid] = rf
    progress(f"Radarr maps: tmdb={len(tmdb_to_id)} title+year={len(ty_to_id)} ids_with_root={len(id_to_root)}")
    return tmdb_to_id, ty_to_id, id_to_root

def build_sonarr_maps(series: List[dict], son_rootfolders: List[dict]):
    progress("Step 5/9: Building Sonarr ID maps…")
    tvdb_to_id, title_to_id = {}, {}
    id_to_root = {}
    roots = [rf.get("path") or rf.get("rootFolderPath") or "" for rf in (son_rootfolders or [])]
    roots = sorted({p.rstrip("/\\") for p in roots if p}, key=len, reverse=True)
    for s in series:
        sid=s.get("id")
        tvdb=s.get("tvdbId")
        title=(s.get("title") or "").strip().lower()
        if isinstance(tvdb,int): tvdb_to_id[tvdb]=sid
        if title: title_to_id[title]=sid
        rf = s.get("rootFolderPath") or ""
        if not rf:
            path = (s.get("path") or "").rstrip("/\\")
            rf = ""
            for r in roots:
                if path.startswith(r):
                    rf = r
                    break
        id_to_root[sid] = rf
    progress(f"Sonarr maps: tvdb={len(tvdb_to_id)} title={len(title_to_id)} ids_with_root={len(id_to_root)}")
    return tvdb_to_id, title_to_id, id_to_root

# ---------- XLSX ----------
def write_xlsx(path: str, sheets: Dict[str, Tuple[List[str], List[dict]]]):
    wb = Workbook()
    default = wb.active
    wb.remove(default)

    for sheet_name, (headers, rows) in sheets.items():
        ws = wb.create_sheet(title=sheet_name[:31])  # Excel tab name limit
        ws.append(headers)
        for row in rows:
            ws.append([row.get(h, "") for h in headers])

        # auto-ish column widths
        for col_idx, h in enumerate(headers, start=1):
            values = [str(row.get(h, "")) for row in rows]
            max_len = max([len(str(h))] + [len(v) for v in values]) if rows else len(str(h))
            max_len = min(max_len, 80)
            ws.column_dimensions[get_column_letter(col_idx)].width = max(10, max_len + 2)

    wb.save(path)
    progress(f"Wrote XLSX -> {path}")

# ---------- utils ----------
def parse_iso_dt(dt_str: str) -> datetime:
    if not dt_str:
        return datetime.max
    s = dt_str.strip()
    # Handle Z and fractional seconds
    if s.endswith("Z"):
        s = s[:-1] + "+00:00"
    try:
        return datetime.fromisoformat(s)
    except ValueError:
        # Trim fractional seconds if too long
        if "." in s:
            head, tail = s.split(".", 1)
            # remove timezone from tail if present
            tz = ""
            if "+" in tail:
                tail, tz = tail.split("+", 1)
                tz = "+" + tz
            elif "-" in tail and tail.count("-") == 2:
                # unlikely, but guard
                pass
            tail = tail[:6]  # microseconds precision
            try:
                return datetime.fromisoformat(head + "." + tail + tz)
            except Exception:
                pass
        return datetime.max

def deduplicate_keep_earliest(rows: List[dict], id_field: str, title_field: str):
    """
    Keep earliest request per (requester, id) key.
    If id missing, fall back to (requester, title.lower()).
    Returns (deduped_rows, duplicates_removed)
    """
    best: Dict[Tuple[str, str], Tuple[datetime, dict]] = {}
    dropped: List[dict] = []

    for row in rows:
        requester = (row.get("requester") or "").strip()
        id_val = row.get(id_field)
        title = (row.get(title_field) or "").strip().lower()

        key_second = str(id_val) if id_val not in (None, "", 0) else title
        key = (requester, key_second)

        d = parse_iso_dt(row.get("requestDate") or "")
        if key not in best:
            best[key] = (d, row)
        else:
            if d < best[key][0]:
                # new earliest; drop previous best
                dropped.append(best[key][1])
                best[key] = (d, row)
            else:
                dropped.append(row)

    deduped = [v[1] for v in best.values()]
    return deduped, dropped

# ---------- main ----------
def main():
    ap = argparse.ArgumentParser(description="Ombi→*arr tag backfill to multi-sheet XLSX (no status filtering) with dedup.")
    ap.add_argument("--write", choices=["radarr","sonarr"], action="append",
                    help="Apply tags to this service (can repeat). Omit for preview only.")
    ap.add_argument("--apply-tags", choices=["add","replace","remove"], default="add",
                    help="Tag apply mode (default add).")
    ap.add_argument("--out", default="ombi_tag_preview.xlsx", help="Output XLSX path.")
    ap.add_argument("--timeout", type=int, default=15)
    ap.add_argument("--retries", type=int, default=3)
    ap.add_argument("--backoff", type=float, default=1.5)
    ap.add_argument("--year-fuzz", type=int, default=1, help="Movie year ± fuzz for title matching.")
    args = ap.parse_args()

    # env check
    req = ["OMBI_URL","OMBI_API_KEY","RADARR_URL","RADARR_API_KEY","SONARR_URL","SONARR_API_KEY"]
    missing=[k for k in req if not os.getenv(k)]
    if missing:
        progress(f"[FATAL] Missing env vars: {', '.join(missing)}"); sys.exit(2)

    http = Http(timeout=args.timeout, max_retries=args.retries, backoff=args.backoff)

    try:
        ombi = Ombi(os.environ["OMBI_URL"], os.environ["OMBI_API_KEY"], http)
        rad  = Radarr(os.environ["RADARR_URL"], os.environ["RADARR_API_KEY"], http)
        son  = Sonarr(os.environ["SONARR_URL"], os.environ["SONARR_API_KEY"], http)

        # Ombi (no status filter; TV per childRequest)
        movies, tv_requests = ombi.fetch_all()

        # Libraries + maps + root folders
        rad_movies = rad.all_movies()
        rad_roots  = rad.root_folders()
        tmdb_to_movie, ty_to_movie, movieid_to_root = build_radarr_maps(rad_movies, rad_roots)

        son_series = son.all_series()
        son_roots  = son.root_folders()
        tvdb_to_series, title_to_series, seriesid_to_root = build_sonarr_maps(son_series, son_roots)

        # Build previews + unmatched
        progress("Step 6/9: Building previews & unmatched…")

        # --- MOVIES (Radarr)
        rad_rows=[]; rad_unmatched=[]
        for i,reqd in enumerate(movies,1):
            if i % 250 == 0: progress(f"… movies processed {i}/{len(movies)}")
            title=(reqd.get("title") or "").strip()
            year=reqd.get("year")
            tmdb=reqd.get("tmdbId")
            requester=reqd.get("requester") or ""
            rdate=reqd.get("requestDate") or ""

            mid=None
            if isinstance(tmdb,int) and tmdb in tmdb_to_movie:
                mid=tmdb_to_movie[tmdb]
            else:
                if title and isinstance(year,int):
                    mid = ty_to_movie.get((title.lower(),year))
                    if mid is None and args.year_fuzz > 0:
                        for dy in range(1, args.year_fuzz+1):
                            mid = ty_to_movie.get((title.lower(), year - dy)) or ty_to_movie.get((title.lower(), year + dy))
                            if mid is not None:
                                break

            root = movieid_to_root.get(mid, "")
            if mid is None:
                rad_unmatched.append({
                    "requestDate": rdate, "requester": requester,
                    "tmdbId": tmdb if isinstance(tmdb,int) else "", "radarrId": "",
                    "title": title, "rootFolder": "", "note": "No Radarr match by tmdbId or (title,year±fuzz)"
                })
            else:
                rad_rows.append({
                    "requestDate": rdate, "requester": requester,
                    "tmdbId": tmdb if isinstance(tmdb,int) else "", "radarrId": mid,
                    "title": title, "rootFolder": root, "note": ""
                })

        # --- TV (Sonarr)
        son_rows=[]; son_unmatched=[]
        for i,reqd in enumerate(tv_requests,1):
            if i % 250 == 0: progress(f"… tv requests processed {i}/{len(tv_requests)}")
            title=(reqd.get("title") or "").strip()
            tvdb=reqd.get("tvdbId")
            requester=reqd.get("requester") or ""
            rdate=reqd.get("requestDate") or ""
            sid=None
            if isinstance(tvdb,int) and tvdb in tvdb_to_series:
                sid=tvdb_to_series[tvdb]
            elif title:
                sid=title_to_series.get(title.lower())
            root = seriesid_to_root.get(sid, "")
            if sid is None:
                son_unmatched.append({
                    "requestDate": rdate, "requester": requester,
                    "tvdbId": tvdb if isinstance(tvdb,int) else "", "sonarrId": "",
                    "title": title, "rootFolder": "", "note": "No Sonarr match by tvdbId or title"
                })
            else:
                son_rows.append({
                    "requestDate": rdate, "requester": requester,
                    "tvdbId": tvdb if isinstance(tvdb,int) else "", "sonarrId": sid,
                    "title": title, "rootFolder": root, "note": ""
                })

        # --- Deduplicate previews (keep earliest per requester+ID/title)
        progress("Step 7/9: Deduplicating previews (keep earliest per requester+ID/title)…")
        rad_rows_dedup, rad_dupes = deduplicate_keep_earliest(rad_rows, "tmdbId", "title")
        son_rows_dedup, son_dupes = deduplicate_keep_earliest(son_rows, "tvdbId", "title")

        # --- Write XLSX (six tabs)
        sheets = {
            "radarr_preview":             (["requestDate","requester","tmdbId","radarrId","title","rootFolder","note"], rad_rows_dedup),
            "sonarr_preview":             (["requestDate","requester","tvdbId","sonarrId","title","rootFolder","note"], son_rows_dedup),
            "radarr_unmatched":           (["requestDate","requester","tmdbId","radarrId","title","rootFolder","note"], rad_unmatched),
            "sonarr_unmatched":           (["requestDate","requester","tvdbId","sonarrId","title","rootFolder","note"], son_unmatched),
            "radarr_duplicates_removed":  (["requestDate","requester","tmdbId","radarrId","title","rootFolder","note"], rad_dupes),
            "sonarr_duplicates_removed":  (["requestDate","requester","tvdbId","sonarrId","title","rootFolder","note"], son_dupes),
        }
        out_path = os.path.abspath(args.out)
        write_xlsx(out_path, sheets)

        progress(f"Preview complete. Unmatched: Radarr={len(rad_unmatched)}, Sonarr={len(son_unmatched)}")
        progress(f"Duplicates removed: Radarr={len(rad_dupes)}, Sonarr={len(son_dupes)}")

        # --- Summary by user (from deduped previews)
        progress("Step 8/9: Summarizing requests by user (deduped previews)…")
        totals = Counter(); movies_by = Counter(); shows_by = Counter()
        for row in rad_rows_dedup: totals[row["requester"] or "[blank]"] += 1; movies_by[row["requester"] or "[blank]"] += 1
        for row in son_rows_dedup: totals[row["requester"] or "[blank]"] += 1; shows_by[row["requester"] or "[blank]"] += 1

        print("\n=== Request counts by user (deduped) ===")
        if totals:
            for user, total in totals.most_common():
                print(f"{user:20} : {total} (movies={movies_by[user]}, shows={shows_by[user]})")
        else:
            print("(no requests)")
        print()

        # --- Apply tags if requested (use deduped previews)
        writes=set(args.write or [])
        if "radarr" in writes:
            progress("Step 9/9: Applying tags to Radarr (deduped)…")
            tag_to_ids: Dict[str,List[int]]={}
            for row in rad_rows_dedup:
                label=row["requester"]; mid=row.get("radarrId")
                if label and isinstance(mid,int): tag_to_ids.setdefault(label,[]).append(int(mid))
            for label, ids in tag_to_ids.items():
                try:
                    tag_id = rad.get_or_create_tag(label)
                    progress(f" Radarr: '{label}' -> {len(ids)} movie(s)")
                    rad.apply_tag(ids, tag_id, mode=args.apply_tags)
                except Exception as e:
                    progress(f"[Radarr] tag apply failed '{label}': {e}")

        if "sonarr" in writes:
            progress("Step 9/9: Applying tags to Sonarr (deduped)…")
            tag_to_ids: Dict[str,List[int]]={}
            for row in son_rows_dedup:
                label=row["requester"]; sid=row.get("sonarrId")
                if label and isinstance(sid,int): tag_to_ids.setdefault(label,[]).append(int(sid))
            for label, ids in tag_to_ids.items():
                try:
                    tag_id = son.get_or_create_tag(label)
                    progress(f" Sonarr: '{label}' -> {len(ids)} series")
                    son.apply_tag(ids, tag_id, mode=args.apply_tags)
                except Exception as e:
                    progress(f"[Sonarr] tag apply failed '{label}': {e}")

        progress("All done.")
        print(f"\nXLSX: {out_path}")
    except Exception as e:
        progress(f"[FATAL] {e}"); sys.exit(1)

if __name__ == "__main__":
    main()
